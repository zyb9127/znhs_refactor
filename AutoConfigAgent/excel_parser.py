#!/usr/bin/env python3
"""
excel_parser.py — Excel 导入模板解析器

支持的 Excel 工作表（Sheet）：
  1. meta         —— Skill 元信息（key/value 两列）
  2. api          —— 接口配置（key/value 两列）
  3. strategy     —— 策略配置（key/value 两列）
  4. templates    —— 话术模板（多列表格）
  5. mock(可选)    —— 接口 Mock 响应（粘贴 JSON 字符串）

最终输出与 code_generator.SkillPackageGenerator 接受的模板字典等价。
"""
from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# 表头说明（用于生成模板与文档）
META_FIELDS: List[Tuple[str, str, str, bool]] = [
    # (key, 说明, 示例, 必填)
    ("province", "省份代码（英文）", "shandong", True),
    ("province_name", "省份中文名", "山东", False),
    ("intent", "业务意图（与 skills-runtime 目录名一致）", "套餐推荐", True),
    ("description", "Skill 功能描述", "山东套餐推荐 Skill", False),
    ("version", "版本号", "1.0.0", False),
    ("author", "作者", "ops-team", False),
]

API_FIELDS: List[Tuple[str, str, str, bool]] = [
    ("name", "接口节点名称（唯一键）", "shandong_package_api", False),
    ("comment", "接口说明", "山东营销推荐接口", False),
    ("url", "接口地址（mock_mode=true 时可留空）", "http://your-api/recommend", False),
    ("method", "HTTP 方法", "POST", False),
    ("headers", "请求头（JSON）", '{"Content-Type":"application/json"}', False),
    ("timeout", "超时（秒）", "30", False),
    ("max_retries", "重试次数", "2", False),
    ("mock_mode", "是否启用 Mock（true/false）", "true", False),
    ("request_template", "请求体模板（JSON）", '{"phone":"{{PHONE}}","intent":"{{INTENT}}"}', False),
    ("response_extract", "响应字段抽取（JSON）", '{"current_package":"bean.mainoffer"}', False),
]

STRATEGY_FIELDS: List[Tuple[str, str, str, bool]] = [
    ("default_strategy", "默认策略（direct/llm）", "direct", False),
    ("top_n", "推荐条数", "3", False),
    ("max_script_length", "话术最大长度", "100", False),
    ("max_parallel_scripts", "并行生成话术数", "3", False),
]

TEMPLATE_COLUMNS: List[Tuple[str, str, str, bool]] = [
    ("template_name", "话术模板名称", "套餐推荐话术", True),
    ("stage", "话术阶段", "切入环节", False),
    ("scene", "场景标识", "", False),
    ("product_id", "产品ID", "", False),
    ("template_content", "话术正文（支持 {pkg_brief} 等占位符）",
     "您好，您当前套餐为{cur_brief}，为您推荐{pkg_brief}", True),
    ("script_requirement", "话术约束（发给 LLM 的指令）", "口语化，80字以内", False),
    ("linked_vars", "关联变量（逗号分隔）", "cur_brief,pkg_brief,diff_str", False),
    ("status", "状态（online/offline）", "online", False),
]


# ───────────────────────── 解析 ─────────────────────────

def _coerce(val: Any) -> Any:
    """把 Excel 单元格转为合适的 Python 值。"""
    if val is None:
        return None
    if isinstance(val, str):
        s = val.strip()
        if s == "":
            return None
        # 数值
        if s.lstrip("-").isdigit():
            try:
                return int(s)
            except Exception:
                return s
        # 布尔
        low = s.lower()
        if low in ("true", "yes"):
            return True
        if low in ("false", "no"):
            return False
        # JSON
        if (s.startswith("{") and s.endswith("}")) or (s.startswith("[") and s.endswith("]")):
            try:
                return json.loads(s)
            except Exception:
                return s
        return s
    return val


def _read_kv_sheet(ws) -> Dict[str, Any]:
    """读取 key/value 两列的工作表。第一行为表头。"""
    out: Dict[str, Any] = {}
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    for row in rows:
        if not row or row[0] in (None, ""):
            continue
        key = str(row[0]).strip()
        val = row[1] if len(row) > 1 else None
        coerced = _coerce(val)
        if coerced is not None:
            out[key] = coerced
    return out


def _read_table_sheet(ws) -> List[Dict[str, Any]]:
    """读取表格型工作表（首行为表头）。"""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h else "" for h in rows[0]]
    out: List[Dict[str, Any]] = []
    for row in rows[1:]:
        if not any(c not in (None, "") for c in row):
            continue
        item: Dict[str, Any] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            v = row[i] if i < len(row) else None
            cv = _coerce(v)
            if cv is not None:
                item[h] = cv
        if item:
            out.append(item)
    return out


def parse_excel(content: bytes) -> Dict[str, Any]:
    """解析 Excel 字节流，返回标准模板字典。"""
    wb = load_workbook(BytesIO(content), data_only=True)
    sheet_names = {n.lower(): n for n in wb.sheetnames}

    def _get(name: str):
        return wb[sheet_names[name]] if name in sheet_names else None

    meta = _read_kv_sheet(_get("meta")) if _get("meta") else {}
    api = _read_kv_sheet(_get("api")) if _get("api") else {}
    strategy = _read_kv_sheet(_get("strategy")) if _get("strategy") else {}
    templates_raw = _read_table_sheet(_get("templates")) if _get("templates") else []

    # mock 工作表（A1 单元格为 JSON 字符串）
    mock_ws = _get("mock")
    if mock_ws is not None:
        cell = mock_ws.cell(row=1, column=1).value
        if cell:
            mock_obj = _coerce(cell)
            if isinstance(mock_obj, (dict, list)):
                api["mock_response"] = mock_obj

    # 模板的 linked_vars 字符串转列表
    templates: List[Dict[str, Any]] = []
    for t in templates_raw:
        item = dict(t)
        lv = item.get("linked_vars")
        if isinstance(lv, str):
            item["linked_vars"] = [s.strip() for s in lv.split(",") if s.strip()]
        templates.append(item)

    return {
        "meta": meta,
        "api": api,
        "strategy": strategy,
        "templates": templates,
    }


# ───────────────────────── 生成模板 Excel ─────────────────────────

_HEADER_FILL = PatternFill("solid", fgColor="0F3460")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_DESC_FILL = PatternFill("solid", fgColor="F0F4FF")
_DESC_FONT = Font(italic=True, color="606266", size=10)
_REQ_FONT = Font(color="C45656", bold=True, size=10)


def _style_header_row(ws, row_idx: int, n_cols: int):
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row_idx, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autosize(ws, widths: List[int]):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_kv_sheet(ws, fields: List[Tuple[str, str, str, bool]], example_row: bool = True):
    """写入 key/value 两列结构。

    布局：
      1: 表头（字段 / 取值 / 说明 / 必填 / 示例）
      2..: 每个字段一行，字段名预填，取值列留空待用户填写
    """
    headers = ["字段", "取值", "说明", "必填", "示例"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    for key, desc, ex, required in fields:
        # 用户实际填写的字段在 B 列；为方便上传，A 列写字段名，B 列留给用户
        ws.append([key, "", desc, "是" if required else "否", ex])
    # 默认填充示例值（用户也可清空重填）
    if example_row:
        for r, (key, desc, ex, required) in enumerate(fields, start=2):
            cell = ws.cell(row=r, column=2, value=ex)
            cell.font = Font(color="909399")
    # 样式
    for r in range(2, 2 + len(fields)):
        ws.cell(row=r, column=3).font = _DESC_FONT
        if ws.cell(row=r, column=4).value == "是":
            ws.cell(row=r, column=4).font = _REQ_FONT
    _autosize(ws, [22, 36, 50, 8, 36])


def _write_templates_sheet(ws):
    headers = [c[0] for c in TEMPLATE_COLUMNS]
    descs = [c[1] for c in TEMPLATE_COLUMNS]
    examples = [c[2] for c in TEMPLATE_COLUMNS]

    ws.append(headers)
    _style_header_row(ws, 1, len(headers))

    # 第二行：字段说明（淡色）
    ws.append(descs)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=2, column=c)
        cell.fill = _DESC_FILL
        cell.font = _DESC_FONT
        cell.alignment = Alignment(wrap_text=True, vertical="center")

    # 第三行：示例
    ws.append(examples)
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=3, column=c)
        cell.font = Font(color="909399")
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    # 第二个示例（不同 stage）
    second = [
        "切入环节话术", "切入环节", "", "",
        "您好，看到您近期流量使用较多，为您推荐{pkg_brief}",
        "口语化，50字以内", "pkg_brief,usage_line", "online",
    ]
    ws.append(second)
    for c in range(1, len(headers) + 1):
        ws.cell(row=4, column=c).font = Font(color="909399")

    _autosize(ws, [18, 12, 10, 10, 50, 30, 26, 10])
    ws.row_dimensions[2].height = 36
    ws.row_dimensions[3].height = 48


def _write_readme_sheet(ws):
    ws.append(["自动配置智能体 · Excel 导入模板说明"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14, color="0F3460")

    rows = [
        [""],
        ["本模板共包含 5 个工作表，请按需填写后上传："],
        [""],
        ["工作表", "用途", "填写方式"],
        ["meta", "Skill 元信息（省份/意图等）", "B 列填写对应取值"],
        ["api", "接口配置（URL/请求/响应字段）", "B 列填写；JSON 字段直接粘贴 JSON 字符串"],
        ["strategy", "推荐策略配置", "B 列填写；可全部留空使用默认值"],
        ["templates", "话术模板（每行一个模板）", "从第 4 行起填写，可增加多行"],
        ["mock", "接口 Mock 响应（可选）", "在 A1 单元格粘贴 JSON 字符串"],
        [""],
        ["注意事项："],
        ["1. 必填项标 “是” 的字段必须填写，否则导入会失败。"],
        ["2. JSON 字段必须为合法 JSON（双引号包裹），可使用工具校验。"],
        ["3. templates 工作表中第 2、3 行为示例和说明，导入时会跳过空模板名的行。"],
        ["4. linked_vars 列填写英文逗号分隔的变量名，如 cur_brief,pkg_brief。"],
        ["5. 建议先填写 meta + templates，api 留空配合 mock_mode=true 测试。"],
    ]
    for r in rows:
        ws.append(r)

    # 表头美化
    _style_header_row(ws, 4, 3)
    _autosize(ws, [18, 36, 50])


def build_template_workbook() -> bytes:
    """生成 Excel 导入模板（字节流）。"""
    wb = Workbook()
    # 默认 sheet 改名为 readme
    readme = wb.active
    readme.title = "readme"
    _write_readme_sheet(readme)

    meta_ws = wb.create_sheet("meta")
    _write_kv_sheet(meta_ws, META_FIELDS)

    api_ws = wb.create_sheet("api")
    _write_kv_sheet(api_ws, API_FIELDS)

    strategy_ws = wb.create_sheet("strategy")
    _write_kv_sheet(strategy_ws, STRATEGY_FIELDS)

    tpl_ws = wb.create_sheet("templates")
    _write_templates_sheet(tpl_ws)

    mock_ws = wb.create_sheet("mock")
    mock_ws["A1"] = (
        '{"rtnCode":"0","bean":{"mainoffer":{"offerName":"99元畅享套餐","initFee":99},'
        '"recommend_results":[{"rank":1,"offerName":"129元5G套餐","initFee":129}]}}'
    )
    mock_ws["A1"].alignment = Alignment(wrap_text=True, vertical="top")
    _autosize(mock_ws, [120])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def get_template_doc() -> Dict[str, Any]:
    """返回 Excel 模板的字段说明文档（JSON 形式，供前端展示）。"""
    def _to_doc(fields: List[Tuple[str, str, str, bool]]):
        return [
            {"field": k, "desc": d, "example": ex, "required": req}
            for k, d, ex, req in fields
        ]

    return {
        "sheets": [
            {"name": "meta", "title": "Skill 元信息", "type": "kv", "fields": _to_doc(META_FIELDS)},
            {"name": "api", "title": "接口配置", "type": "kv", "fields": _to_doc(API_FIELDS)},
            {"name": "strategy", "title": "策略配置", "type": "kv", "fields": _to_doc(STRATEGY_FIELDS)},
            {"name": "templates", "title": "话术模板", "type": "table", "fields": _to_doc(TEMPLATE_COLUMNS)},
            {"name": "mock", "title": "接口 Mock 响应（可选）", "type": "json",
             "fields": [{"field": "A1 单元格", "desc": "粘贴接口的 Mock 响应 JSON",
                         "example": '{"bean":{...}}', "required": False}]},
        ]
    }
